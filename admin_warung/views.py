<<<<<<< HEAD
from os import name
from django.http import response, HttpResponseNotAllowed, HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy

from .models import Barang
from .barang_form import BarangForm, ExcelUploadForm
=======
from django.http import response, HttpResponseNotAllowed
from .models import *
from admin_warung.models import Barang
from .barang_form import BarangForm, ExcelUploadForm
from django.shortcuts import redirect, render, get_object_or_404
>>>>>>> 5053ecd1431c1efc60edec2a8816e4c30467af27

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill

<<<<<<< HEAD
import copy
from datetime import datetime
=======
from transaksi.models import Transaksi
from keranjang.models import ItemKeranjang

>>>>>>> 5053ecd1431c1efc60edec2a8816e4c30467af27

def dashboard(request):
    barangs = Barang.objects.all()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'konfirmasi':
            idTransaksi = request.POST.get('transaksi')
            update_transaksi(idTransaksi)

            username = request.user.username
            set_barang_terjual(username, idTransaksi)

            redirect('admin_warung:dashboard')
        else:
            idTransaksi = request.POST.get('transaksi')
            update_status(idTransaksi)

            redirect('admin_warung:dashboard')

    context = {
        "excel_upload_form": ExcelUploadForm(),
        "barangs": barangs,
    }

    context['transaksi'] = get_all_transaksi()

    # print(context)
    return render(request, 'dashboard.html', context=context)

def detail_barang_admin_view(request, id):
    response = {}
    barang = get_object_or_404(Barang, id=id)
    response['barang'] = barang

    return render(request, 'detail_barang_admin.html', response)


def hapus_barang_view(request, id):
    response = {}
    barang = get_object_or_404(Barang, id=id)
    response['barang'] = barang

    if(request.method == 'POST'):
        barang.delete()

        return redirect('/dashboard')

    return render(request, 'hapus_barang.html', response)


def tambah_barang_view(request):
    response = {}
    if(request.method == 'POST'):
        nama = request.POST.get("nama")
        harga = request.POST.get("harga")
        stok = request.POST.get("stok")
        terjual = request.POST.get("terjual")
        deskripsi = request.POST.get("deskripsi")
        image_url = request.POST.get("image_url")

        barang = Barang.objects.create(nama=nama, harga=harga, stok=stok, terjual=terjual, deskripsi=deskripsi, image_url=image_url)
        redir = f'/dashboard/detail-barang/{barang.id}'
        return redirect(redir)
    
    return render(request, 'tambah_barang.html', response)


def ubah_barang_view(request, id):
    response = {}
    barang = get_object_or_404(Barang, id=id)
    response['barang'] = barang

    if(request.method == 'POST'):
        nama = request.POST.get("nama")
        harga = request.POST.get("harga")
        stok = request.POST.get("stok")
        terjual = request.POST.get("terjual")
        deskripsi = request.POST.get("deskripsi")
        image_url = request.POST.get("image_url")

        Barang.objects.filter(id=id).update(nama=nama, harga=harga, stok=stok, terjual=terjual, deskripsi=deskripsi, image_url=image_url)
        redir = f'/dashboard/detail-barang/{id}'
        return redirect(redir)

    return render(request, 'ubah_barang.html', response)


def list_ubah_view(request):
    response = {}
    barangs = Barang.objects.all()
    response['barangs'] = barangs
    return render(request, 'list_ubah.html', response)

def list_hapus_view(request):
    response = {}
    barangs = Barang.objects.all()
    response['barangs'] = barangs
    return render(request, 'list_hapus.html', response)

# def import_barang(request):
#     if request.method == "GET":
#         context = {
#             "excel_upload_form": ExcelUploadForm(),
#         }
#         return render(request, 'dashboard.html',)

#     if request.method == "POST":
#         context = {"excel_upload_form": ExcelUploadForm()}
#         return render(request, 'dashboard.html',)

#     return HttpResponseNotAllowed(["GET", "POST"])


def import_barang(request):
    form = ExcelUploadForm(request.POST, request.FILES)
    if form.is_valid():
        workbook = load_workbook(request.FILES["excel_file"])
        sheet = workbook.active
        Barang.objects.all().delete()   
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if(row[0] == None):
                break
            barang = Barang(nama=row[1],
                            harga=row[2],
                            stok=row[3],
                            terjual=row[4],
                            deskripsi=row[5],
                            image_url=row[6])

            barang.save()
            
        return redirect(reverse_lazy('admin_warung:dashboard'))
    

def export_barang():
    wb = Workbook()

    ws = wb.active
    ws.title = "Barang"
    
    # Creating column header
    title_list = ['No.', 'Nama', 'Harga', 'Stok', 'Terjual', 'Deskripsi', 'Image URL']
    ws.append(title_list)
    
    # Data insertion
    all_barang = Barang.objects.all().order_by('id')
    num_of_data = len(all_barang)
    
    for barang in all_barang:
        ws.append(barang.get_all_value())
        

<<<<<<< HEAD
    ### Styling ###
    
    # Column Size
    ws.column_dimensions['A'].width = 6
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 25
        if col == 6:
            ws.column_dimensions[get_column_letter(col)].width = 50
    
    # Row Size
    ws.row_dimensions[1].height = 40

    for row in range(2, num_of_data + 2):
        ws.row_dimensions[row].height = 60

    # Coloring & format
    
    for row in range(1, num_of_data + 2):
        for col in range(1, 8):
            ws["{}{}".format(get_column_letter(col), row)].alignment = Alignment(
                horizontal="center", vertical="center")

            ws[f'{get_column_letter(col)}{row}'].border = Border(left=Side(border_style="thin",
                                                                            color='00000000'),
                                                                  right=Side(border_style="thin",
                                                                             color='00000000'),
                                                                  top=Side(border_style="thin",
                                                                           color='00000000'),
                                                                  bottom=Side(border_style="thin",
                                                                              color='00000000'),
                                                                  )
            if(row == 1):
                ws[f'{get_column_letter(col)}{row}'].fill = PatternFill( "solid", fgColor="0B3C5C")
                
    for row in ws.iter_rows():
        for cell in row:
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            cell.alignment = alignment

    return wb

def download(_):
    response = HttpResponse(save_virtual_workbook(
        export_barang()), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="Export_Barang-{datetime.now()}.xlsx"'
    return response



    
=======
    return HttpResponseNotAllowed(["GET", "POST"])



def get_all_transaksi():
    context = []

    transaksi = Transaksi.objects.all()

    for tr in transaksi:
        id = tr.idTransaksi
        date = tr.timeStamp
        pembeli = tr.usernamePembeli
        total = tr.totalPesanan
        cara = tr.caraPembayaran
        status = tr.statusTransaksi
        context.append([id, date, pembeli, total, cara, status])

    return context

def update_transaksi(id):
    context = []

    transaksi = Transaksi.objects.get(idTransaksi = id)

    transaksi.statusTransaksi = 'transaksi berhasil'

    transaksi.save()

    return context

def update_status(id):
    context = []

    transaksi = Transaksi.objects.get(idTransaksi = id)

    transaksi.statusTransaksi = 'menunggu konfirmasi pembayaran'

    transaksi.save()

    return context

def set_barang_terjual(username, id):
    tr = Transaksi.objects.get(idTransaksi = id)
    item_keranjang = ItemKeranjang.objects.filter(pelanggan__username = username,  transaksi = tr)

    for item in item_keranjang:
        item = ItemKeranjang.objects.get(pelanggan = item.pelanggan, barang = item.barang, transaksi = tr)
        barang = item.barang
        barang.terjual = barang.terjual + item.jumlah_item
        barang.save()
>>>>>>> 5053ecd1431c1efc60edec2a8816e4c30467af27
